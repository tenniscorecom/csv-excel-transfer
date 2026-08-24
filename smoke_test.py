"""一時フォルダだけを使う実走行確認。"""

import tempfile
from pathlib import Path
from unittest.mock import patch

from comken import Config, dry_run
from openpyxl import Workbook, load_workbook

from src.run import run


def _prepare(root: Path) -> Config:
    csv_folder, input_folder, output_folder = root / "csv", root / "input", root / "output"
    csv_folder.mkdir()
    input_folder.mkdir()
    (csv_folder / "west.csv").write_text("お客様ID,お名前\nC001,山田一郎\n", encoding="utf-8")
    (csv_folder / "east.csv").write_text("お客様ID,お名前\nC002,鈴木三郎\n", encoding="utf-8")
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet.append(["業務用ID", "氏名", "備考"])
    sheet.append(["C001", "", "保持"])
    workbook.save(input_folder / "input.xlsx")
    workbook.close()
    config_path = root / "config.ini"
    config_path.write_text(
        f"[FILES]\nOUTPUT_EXCEL_FOLDER = {output_folder}\n"
        f"INPUT_EXCEL_FOLDER = {input_folder}\nINPUT_CSV_FOLDER = {csv_folder}\n\n"
        "[EXCEL]\nINPUT_NAME = input.xlsx\nOUTPUT_PREFIX = 最終_\n"
        "READ_PASSWORD =\nWRITE_PASSWORD =\n\n"
        "[CSV]\nWEST = west.csv\nEAST = east.csv\n\n[TRANSFER_MAPPING]\nお名前 = 氏名\n",
        encoding="utf-8",
    )
    return Config(config_path)


def main() -> None:
    with tempfile.TemporaryDirectory(prefix="csv_excel_transfer_") as directory:
        root = Path(directory)
        result = run(_prepare(root))
        workbook = load_workbook(result.output_path)
        assert workbook["Sheet1"]["B2"].value == "山田一郎"
        assert workbook["Sheet1"]["C2"].value == "保持"
        workbook.close()

    with tempfile.TemporaryDirectory(prefix="csv_excel_transfer_dry_") as directory:
        root = Path(directory)
        settings = _prepare(root)
        with dry_run():
            result = run(settings)
        assert not result.output_path.exists()

    with tempfile.TemporaryDirectory(prefix="csv_excel_transfer_password_") as directory:
        root = Path(directory)
        settings = _prepare(root)
        settings.EXCEL.READ_PASSWORD = "read-password"
        settings.EXCEL.WRITE_PASSWORD = "write-password"
        # パスワード保存は COM の ExcelCOMHandler.save_as で行う。
        # COM を起動しないよう __init__ / close を no-op にして Excel プロセスを起こさず、
        # save_as だけ Mock に差し替えて呼び出し内容を検証する。
        with patch("src.run.ExcelCOMHandler.__init__", return_value=None), patch(
            "src.run.ExcelCOMHandler.close", return_value=None
        ), patch("src.run.ExcelCOMHandler.save_as", autospec=True) as save:
            result = run(settings)
        assert save.call_count == 1
        _, saved_path = save.call_args.args
        assert saved_path == result.output_path
        assert save.call_args.kwargs == {
            "read_pw": "read-password",
            "write_pw": "write-password",
        }


if __name__ == "__main__":
    main()