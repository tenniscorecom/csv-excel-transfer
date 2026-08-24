"""run.py の業務フロー全体。"""

from pathlib import Path
from unittest.mock import patch

import pytest
from comken import dry_run
from comken.exceptions import (
    ComkenError,
    ExcelColumnNotFoundError,
    TableDuplicateKeyError,
)
from openpyxl import load_workbook

from src.run import _paths, run
from tests.conftest import SAMPLE_MAPPING


def _write_config(
    tmp_path: Path,
    *,
    mapping: dict[str, str] | None = None,
    read_password: str = "",
    write_password: str = "",
) -> Path:
    mapping = SAMPLE_MAPPING if mapping is None else mapping
    lines = [
        "[FILES]",
        f"OUTPUT_EXCEL_FOLDER = {tmp_path / 'output'}",
        f"INPUT_EXCEL_FOLDER = {tmp_path}",
        f"INPUT_CSV_FOLDER = {tmp_path}",
        "",
        "[EXCEL]",
        "INPUT_NAME = input.xlsx",
        "OUTPUT_PREFIX = 最終_",
        f"READ_PASSWORD = {read_password}",
        f"WRITE_PASSWORD = {write_password}",
        "",
        "[CSV]",
        "WEST = west.csv",
        "EAST = east.csv",
        "",
        "[TRANSFER_MAPPING]",
        *(f"{source} = {destination}" for source, destination in mapping.items()),
    ]
    config_path = tmp_path / "config.ini"
    config_path.write_text("\n".join(lines), encoding="utf-8")
    return config_path


def _files(tmp_path: Path, make_csv, make_input_book) -> tuple[Path, Path, Path]:
    west = make_csv(
        tmp_path / "west.csv",
        ["お客様ID", "お名前", "ご住所", "電話番号"],
        [("C001", "山田一郎", "大阪", "06-0000")],
    )
    east = make_csv(
        tmp_path / "east.csv",
        ["お客様ID", "お名前", "ご住所", "電話番号"],
        [("C002", "鈴木三郎", "東京", "03-0000")],
    )
    book = make_input_book(
        tmp_path / "input.xlsx",
        [("C001", "旧氏名", "旧住所", "旧電話"), ("UNKNOWN", "維持", "維持", "維持")],
    )
    return west, east, book


def test_run_transfers_preserves_unmatched_and_deletes_sources(
    tmp_path: Path, make_csv, make_input_book, use_config
) -> None:
    west, east, book = _files(tmp_path, make_csv, make_input_book)
    use_config(_write_config(tmp_path))
    output = tmp_path / "output" / "最終_input.xlsx"
    assert run() == output
    workbook = load_workbook(output)
    rows = list(workbook["Sheet1"].values)
    workbook.close()
    assert rows[1] == ("C001", "山田一郎", "大阪", "06-0000")
    assert rows[2] == ("UNKNOWN", "維持", "維持", "維持")
    assert not west.exists() and not east.exists() and not book.exists()


def test_run_keeps_all_input_columns(
    tmp_path: Path, make_csv, make_input_book, use_config
) -> None:
    _files(tmp_path, make_csv, make_input_book)
    workbook = load_workbook(tmp_path / "input.xlsx")
    sheet = workbook["Sheet1"]
    sheet["E1"], sheet["E2"] = "備考", "残す"
    workbook.save(tmp_path / "input.xlsx")
    workbook.close()
    use_config(_write_config(tmp_path))
    run()
    workbook = load_workbook(tmp_path / "output" / "最終_input.xlsx")
    assert workbook["Sheet1"]["E2"].value == "残す"
    workbook.close()


def test_run_rejects_cross_file_duplicate(
    tmp_path: Path, make_csv, make_input_book, use_config
) -> None:
    west, east, book = _files(tmp_path, make_csv, make_input_book)
    make_csv(east, ["お客様ID", "お名前", "ご住所", "電話番号"], [("C001", "重複", "東京", "03")])
    use_config(_write_config(tmp_path))
    with pytest.raises(TableDuplicateKeyError, match="お客様ID"):
        run()
    assert west.exists() and east.exists() and book.exists()


def test_run_rejects_empty_csv(
    tmp_path: Path, make_csv, make_input_book, use_config
) -> None:
    _files(tmp_path, make_csv, make_input_book)
    make_csv(tmp_path / "east.csv", ["お客様ID", "お名前", "ご住所", "電話番号"], [])
    use_config(_write_config(tmp_path))
    with pytest.raises(ComkenError, match="データ行がありません"):
        run()


def test_run_rejects_missing_excel_column(
    tmp_path: Path, make_csv, make_input_book, use_config
) -> None:
    west, east, book = _files(tmp_path, make_csv, make_input_book)
    use_config(_write_config(tmp_path, mapping={"お名前": "存在しない列"}))
    with pytest.raises(ExcelColumnNotFoundError):
        run()
    assert west.exists() and east.exists() and book.exists()


def test_run_dry_run_does_not_write_or_delete(
    tmp_path: Path, make_csv, make_input_book, use_config
) -> None:
    west, east, book = _files(tmp_path, make_csv, make_input_book)
    use_config(_write_config(tmp_path))
    with dry_run():
        run()
    assert west.exists() and east.exists() and book.exists()
    assert not (tmp_path / "output" / "最終_input.xlsx").exists()


def test_run_passes_both_passwords_to_save_without_starting_com(
    tmp_path: Path, make_csv, make_input_book, use_config
) -> None:
    _files(tmp_path, make_csv, make_input_book)
    output = tmp_path / "output" / "最終_input.xlsx"
    use_config(
        _write_config(
            tmp_path,
            read_password="read-secret",
            write_password="write-secret",
        )
    )

    # パスワード保存は COM の ExcelCOMHandler.save_as で行う。
    # COM を起動しないよう __init__ / close を no-op にして Excel プロセスを起こさず、
    # save_as だけ Mock に差し替えて呼び出し内容を検証する。
    with patch("src.run.ExcelCOMHandler.__init__", return_value=None), patch(
        "src.run.ExcelCOMHandler.close", return_value=None
    ), patch("src.run.ExcelCOMHandler.save_as", autospec=True) as save:
        assert run() == output

    assert save.call_count == 1
    _, saved_path = save.call_args.args
    assert saved_path == output
    assert save.call_args.kwargs == {
        "read_pw": "read-secret",
        "write_pw": "write-secret",
    }


def test_paths_rejects_output_that_overwrites_input_csv(
    tmp_path: Path, use_config
) -> None:
    config_path = _write_config(tmp_path)
    test_config = use_config(config_path)
    test_config.FILES.OUTPUT_EXCEL_FOLDER = test_config.FILES.INPUT_CSV_FOLDER
    test_config.CSV.WEST = "最終_input.xlsx"

    with pytest.raises(ComkenError):
        _paths()


def test_paths_rejects_same_file_for_both_csv_inputs(
    tmp_path: Path, use_config
) -> None:
    config_path = _write_config(tmp_path)
    test_config = use_config(config_path)
    test_config.CSV.EAST = test_config.CSV.WEST

    with pytest.raises(ComkenError):
        _paths()
