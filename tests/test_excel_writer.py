"""tests/test_excel_writer.py — INPUT エクセルへの転記と保存"""

from pathlib import Path
from unittest.mock import MagicMock

import pytest
from openpyxl import load_workbook

import src.excel_writer as excel_writer
from src.csv_lookup import merge_lookups
from src.excel_writer import transfer_and_save
from tests.conftest import SAMPLE_MAPPING


def _lookup_for(tmp_path: Path, make_csv) -> dict[str, dict[str, str]]:
    west = make_csv(
        tmp_path / "west.csv",
        ["お客様ID", "お名前", "ご住所", "電話番号"],
        [
            ("C001", "山田一郎", "大阪市", "06-0000-0001"),
            ("C002", "佐藤二郎", "京都市", "075-0000-0002"),
        ],
    )
    east = make_csv(
        tmp_path / "east.csv",
        ["お客様ID", "お名前", "ご住所", "電話番号"],
        [
            ("C003", "鈴木三郎", "東京", "03-0000-0003"),
        ],
    )
    return merge_lookups(west, east, "お客様ID")


def test_transfer_and_save_writes_output_with_expected_name(
    tmp_path: Path, make_csv, make_input_book
) -> None:
    input_path = make_input_book(
        tmp_path / "作業対象.xlsx",
        [("C001", "", "", ""), ("C002", "", "", "")],
    )
    output_path = tmp_path / "最終_作業対象.xlsx"
    lookup = _lookup_for(tmp_path, make_csv)

    matched = transfer_and_save(
        input_path=input_path,
        output_path=output_path,
        sheet_name="Sheet1",
        key_column="業務用ID",
        lookup=lookup,
        mapping=SAMPLE_MAPPING,
        header_row=1,
        password="",
    )

    assert matched == 2
    assert output_path.exists()
    # 元ファイルとは別名で「最終_ + 元のファイル名」になる
    assert output_path.name == "最終_作業対象.xlsx"


def test_password_path_preserves_xlsm_suffix_for_com(
    tmp_path: Path, make_csv, make_input_book, monkeypatch
) -> None:
    input_path = make_input_book(
        tmp_path / "作業対象.xlsm",
        [("C001", "", "", "")],
    )
    output_path = tmp_path / "最終_作業対象.xlsm"
    lookup = _lookup_for(tmp_path, make_csv)

    writer = MagicMock()
    sheet = MagicMock()
    sheet.transfer_by_mapping.return_value = 1
    writer.__enter__.return_value = writer
    writer.sheet.return_value = sheet
    com = MagicMock()
    com.__enter__.return_value = com
    monkeypatch.setattr(excel_writer, "ExcelWriter", MagicMock(return_value=writer))
    monkeypatch.setattr(excel_writer, "ExcelComHandler", MagicMock(return_value=com))

    matched = transfer_and_save(
        input_path=input_path,
        output_path=output_path,
        sheet_name="Sheet1",
        key_column="業務用ID",
        lookup=lookup,
        mapping=SAMPLE_MAPPING,
        header_row=1,
        password="dummy-pw",
    )

    assert matched == 1
    tmp_saved_path = writer.save.call_args.kwargs["path"]
    assert tmp_saved_path.suffix == ".xlsm"
    assert com.save_as.call_args.args[0] == output_path
    # パスワードが COM まで渡されている（秘匿値はログに出さないが、COM には確実に渡す）
    assert com.save_as.call_args.kwargs["read_pw"] == "dummy-pw"
    assert not tmp_saved_path.exists()


def test_password_path_cleans_up_tmp_file_when_com_save_as_fails(
    tmp_path: Path, make_csv, make_input_book, monkeypatch
) -> None:
    """COM の save_as が失敗したときも、openpyxl が出した一時ファイルは必ず消える。"""
    input_path = make_input_book(
        tmp_path / "作業対象.xlsx",
        [("C001", "", "", "")],
    )
    output_path = tmp_path / "最終_作業対象.xlsx"
    lookup = _lookup_for(tmp_path, make_csv)

    # openpyxl 側は実際に tmp_path にファイルを作らせる（実物の unlink を見るため）
    real_writer = MagicMock()

    def _save(path):
        # 実ファイルを作って unlink() で消えることを確認できるようにする
        Path(path).write_text("dummy", encoding="utf-8")

    real_writer.save.side_effect = _save
    real_writer.__enter__.return_value = real_writer
    sheet = MagicMock()
    sheet.transfer_by_mapping.return_value = 1
    real_writer.sheet.return_value = sheet

    # COM は save_as で PermissionError
    com = MagicMock()
    com.__enter__.return_value = com
    com.save_as.side_effect = PermissionError("COM 保存失敗の想定")

    monkeypatch.setattr(excel_writer, "ExcelWriter", MagicMock(return_value=real_writer))
    monkeypatch.setattr(excel_writer, "ExcelComHandler", MagicMock(return_value=com))

    with pytest.raises(PermissionError):
        transfer_and_save(
            input_path=input_path,
            output_path=output_path,
            sheet_name="Sheet1",
            key_column="業務用ID",
            lookup=lookup,
            mapping=SAMPLE_MAPPING,
            header_row=1,
            password="dummy-pw",
        )

    # tmp_path 配下に ".最終_作業対象.xlsx.*" 形式の一時ファイルが残っていない
    # パスワード経路では prefix=.{output_path.name}. で `.最終_作業対象.xlsx.<rand>.xlsx` ができる
    leaked = list(tmp_path.glob(".最終_作業対象.xlsx.*"))
    assert leaked == [], f"COM 失敗後に一時ファイルが残っています: {leaked}"


def test_transfer_and_save_actually_writes_transferred_values(
    tmp_path: Path, make_csv, make_input_book
) -> None:
    input_path = make_input_book(
        tmp_path / "作業対象.xlsx",
        [("C003", "", "", "")],
    )
    output_path = tmp_path / "最終_作業対象.xlsx"
    lookup = _lookup_for(tmp_path, make_csv)

    transfer_and_save(
        input_path=input_path,
        output_path=output_path,
        sheet_name="Sheet1",
        key_column="業務用ID",
        lookup=lookup,
        mapping=SAMPLE_MAPPING,
        header_row=1,
        password="",
    )

    workbook = load_workbook(output_path)
    sheet = workbook["Sheet1"]
    # 1行目=ヘッダー、2行目=データ（C003 = 鈴木三郎）
    assert sheet["B2"].value == "鈴木三郎"
    assert sheet["C2"].value == "東京"
    assert sheet["D2"].value == "03-0000-0003"
    workbook.close()


def test_transfer_and_save_skips_rows_with_unknown_key(
    tmp_path: Path, make_csv, make_input_book
) -> None:
    input_path = make_input_book(
        tmp_path / "作業対象.xlsx",
        [("C001", "", "", ""), ("UNKNOWN", "", "", "")],
    )
    output_path = tmp_path / "最終_作業対象.xlsx"
    lookup = _lookup_for(tmp_path, make_csv)

    matched = transfer_and_save(
        input_path=input_path,
        output_path=output_path,
        sheet_name="Sheet1",
        key_column="業務用ID",
        lookup=lookup,
        mapping=SAMPLE_MAPPING,
        header_row=1,
        password="",
    )

    # C001 だけがヒットする。UNKNOWN はスキップされる
    assert matched == 1
    workbook = load_workbook(output_path)
    sheet = workbook["Sheet1"]
    assert sheet["B2"].value == "山田一郎"
    # UNKNOWN 行は touch されない（既存の None のまま）
    assert sheet["B3"].value is None
    workbook.close()
